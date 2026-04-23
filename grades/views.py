from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.db import transaction
from academics.models import CourseOffering, Semester
from grades.models import MarksSubmission, FinalGrade, SemesterResult
from accounts.utils import is_faculty, is_student, is_exam_section
import openpyxl

User = get_user_model()

GRADE_POINTS = {
    "EX" : 10,
    "A" : 9,
    "B" : 8,
    "C" : 7,
    "D" : 6,
    "P" : 5,
    "M" : 4,
    "F" : 0,
    "X" : 0,
    "R" : 0
}
# =========================================================
# 🧑‍🏫 FACULTY MARKS UPLOAD
# =========================================================
@login_required
def faculty_submit_marks(request, offering_id):

    if not is_faculty(request.user):
        return HttpResponseForbidden()

    offering = get_object_or_404(CourseOffering, id=offering_id)

    students = User.objects.filter(
        enrollment__offering=offering,
        enrollment__is_active=True
    )

    student_map = {s.username: s for s in students}

    submissions_qs = MarksSubmission.objects.filter(course_offering=offering)
    submission_map = {m.student_id: m for m in submissions_qs}

    # 🚫 Block if already locked
    if submissions_qs.filter(is_locked=True).exists():
        return render(request, "faculty/submit_marks.html", {
            "offering": offering,
            "students": students,
            "submissions": submission_map,
            "errors": ["Marks are locked. No further changes allowed."]
        })

    if request.method == "POST":

        # =========================
        # 🔒 LOCK MARKS
        # =========================
        if "lock_marks" in request.POST:

            total_students = students.count()
            submitted_count = submissions_qs.count()

            if submitted_count != total_students:
                return render(request, "faculty/submit_marks.html", {
                    "offering": offering,
                    "students": students,
                    "submissions": submission_map,
                    "errors": ["Cannot lock: All students must have marks."]
                })

            incomplete = submissions_qs.filter(
                minor1__isnull=True
            ) | submissions_qs.filter(
                minor2__isnull=True
            ) | submissions_qs.filter(
                mid__isnull=True
            ) | submissions_qs.filter(
                end__isnull=True
            )

            if incomplete.exists():
                return render(request, "faculty/submit_marks.html", {
                    "offering": offering,
                    "students": students,
                    "submissions": submission_map,
                    "errors": ["Cannot lock: Some marks are missing."]
                })

            submissions_qs.update(is_locked=True)
            return redirect("grades:faculty_submit_marks", offering_id=offering.id)

        # =========================
        # 📥 EXCEL UPLOAD
        # =========================
        file = request.FILES.get("file")

        if not file:
            return render(request, "faculty/submit_marks.html", {
                "offering": offering,
                "students": students,
                "submissions": submission_map,
                "errors": ["No file uploaded."]
            })

        if not file.name.endswith(".xlsx"):
            return render(request, "faculty/submit_marks.html", {
                "errors": ["Only .xlsx files allowed."]
            })

        try:
            wb = openpyxl.load_workbook(file)
        except:
            return render(request, "faculty/submit_marks.html", {
                "errors": ["Invalid Excel file."]
            })

        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        expected = ["Roll No", "Minor1", "Minor2", "Mid", "End"]

        if headers != expected:
            return render(request, "faculty/submit_marks.html", {
                "offering": offering,
                "errors": [f"Invalid format. Expected: {expected}"]
            })

        errors = []
        seen_rolls = set()
        new_submissions = []
        updated_submissions = []

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            roll_no, minor1, minor2, mid, end = row

            roll_no = str(roll_no).strip() if roll_no else None

            if not roll_no:
                errors.append(f"Row {row_index}: Missing Roll No")
                continue

            if roll_no in seen_rolls:
                errors.append(f"Row {row_index}: Duplicate Roll No")
                continue

            seen_rolls.add(roll_no)

            student = student_map.get(roll_no)
            if not student:
                errors.append(f"Row {row_index}: Student not found")
                continue

            submission = submission_map.get(student.id)

            def validate(value, field):
                if value in [None, "", "None"]:
                    return None
                try:
                    value = float(value)
                    if value < 0 or value > 100:
                        raise ValueError
                    return value
                except:
                    errors.append(f"Row {row_index}: Invalid {field}")
                    return None

            minor1 = validate(minor1, "Minor1")
            minor2 = validate(minor2, "Minor2")
            mid = validate(mid, "Mid")
            end = validate(end, "End")

            if submission:
                submission.minor1 = minor1
                submission.minor2 = minor2
                submission.mid = mid
                submission.end = end
                submission.submitted_by = request.user
                updated_submissions.append(submission)
            else:
                new_submissions.append(
                    MarksSubmission(
                        student=student,
                        course_offering=offering,
                        minor1=minor1,
                        minor2=minor2,
                        mid=mid,
                        end=end,
                        submitted_by=request.user
                    )
                )

        if len(seen_rolls) != students.count():
            errors.append("File must contain ALL students exactly once.")

        if errors:
            return render(request, "faculty/submit_marks.html", {
                "errors": errors,
                "students": students,
                "submissions": submission_map,
                "offering": offering
            })

        with transaction.atomic():
            if new_submissions:
                MarksSubmission.objects.bulk_create(new_submissions)
            if updated_submissions:
                MarksSubmission.objects.bulk_update(
                    updated_submissions,
                    ["minor1", "minor2", "mid", "end", "submitted_by"]
                )

        return redirect("grades:faculty_submit_marks", offering_id=offering.id)

    return render(request, "faculty/submit_marks.html", {
        "offering": offering,
        "students": students,
        "submissions": submission_map,
    })


# =========================================================
# 🏫 EXAM UPLOAD FINAL GRADES
# =========================================================
import openpyxl
from django.db import transaction

@login_required
def exam_upload_final_grades(request, offering_id):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    offering = get_object_or_404(CourseOffering, id=offering_id)

    # 🔒 Block if frozen
    if FinalGrade.objects.filter(course_offering=offering, is_frozen=True).exists():
        return HttpResponseForbidden("Already frozen, cannot update")

    submissions = MarksSubmission.objects.filter(course_offering=offering)

    if not submissions.exists():
        return HttpResponseForbidden("No marks uploaded")

    if submissions.filter(is_locked=False).exists():
        return HttpResponseForbidden("Marks must be locked before grade upload")

    preview_data = []
    errors = []

    valid_grades = ["EX", "A", "B", "C", "D", "P", "M", "F", "X", "R"]

    # =========================
    # HANDLE POST
    # =========================
    if request.method == "POST":

        file = request.FILES.get("file")

        if not file:
            return render(request, "exam/upload_final_grades.html", {
                "error": "Upload file required",
                "offering": offering
            })

        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        expected = ["Roll No", "Name", "Course Code", "Course Name", "Grade"]

        if headers != expected:
            return render(request, "exam/upload_final_grades.html", {
                "error": f"Invalid format. Expected {expected}",
                "offering": offering
            })

        processed_students = set()

        # 🔥 PARSE FILE (FOR PREVIEW OR SAVE)
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):

            roll_no, name, _, _, grade = row
            roll_no = str(roll_no).strip()

            try:
                student = User.objects.filter(username__iexact=roll_no).first()

                if not student:
                    errors.append(f"Row {row_index}: Student not found ({roll_no})")
                    continue
            except User.DoesNotExist:
                errors.append(f"Row {row_index}: Student not found")
                continue

            if not submissions.filter(student=student).exists():
                errors.append(f"Row {row_index}: Not in this course")
                continue

            grade = str(grade).strip().upper()

            if grade not in valid_grades:
                errors.append(f"Row {row_index}: Invalid grade")
                continue

            processed_students.add(student.id)

            preview_data.append({
                "roll_no": roll_no,
                "name": name,
                "student": student,
                "grade": grade
            })

        # =========================
        # PREVIEW ONLY
        # =========================
        if "preview" in request.POST:
            return render(request, "exam/upload_final_grades.html", {
                "offering": offering,
                "preview_data": preview_data,
                "errors": errors
            })

        # =========================
        # CONFIRM & SAVE
        # =========================
        if "confirm" in request.POST:

            if errors:
                return render(request, "exam/upload_final_grades.html", {
                    "offering": offering,
                    "preview_data": preview_data,
                    "errors": errors
                })

            # 🔥 Validate all students covered
            if len(processed_students) != submissions.count():
                return render(request, "exam/upload_final_grades.html", {
                    "error": "File must contain ALL students",
                    "offering": offering
                })

            with transaction.atomic():

                for row in preview_data:

                    FinalGrade.objects.update_or_create(
                        student=row["student"],
                        course=offering.course,
                        course_offering=offering,
                        defaults={
                            "grade_letter": row["grade"],
                            "published_by": request.user
                        }
                    )

            return redirect("exam_dashboard")

    # =========================
    # GET REQUEST
    # =========================
    return render(request, "exam/upload_final_grades.html", {
        "offering": offering
    })

# =========================================================
# 🔒 FREEZE RESULTS
# =========================================================
@login_required
def exam_freeze_results(request, offering_id):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    offering = get_object_or_404(CourseOffering, id=offering_id)

    total_students = User.objects.filter(role='STUDENT').count()
    

    submissions = MarksSubmission.objects.filter(course_offering=offering)
    grades = FinalGrade.objects.filter(course_offering=offering)

    # ❌ No marks
    if not submissions.exists():
        return HttpResponseForbidden("No students found")

    # ❌ Not all grades uploaded
    if grades.count() != submissions.count():
        return HttpResponseForbidden("Grades are incomplete")

    # ❌ Already frozen
    if grades.filter(is_frozen=True).exists():
        return HttpResponseForbidden("Already frozen")

    # ✅ Freeze all
    grades.update(is_frozen=True)

    if request.method == "POST":
        grades.update(is_frozen=True)
        return redirect("dashboard")

    return render(request, "exam/confirm_freeze.html", {
        "offering": offering
    })


# =========================================================
# 📊 DASHBOARD
# =========================================================
@login_required
def exam_dashboard(request):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    offerings = CourseOffering.objects.select_related("course", "semester")

    dashboard_data = []

    for offering in offerings:

        submissions = MarksSubmission.objects.filter(course_offering=offering)

        total_students = User.objects.filter(role='STUDENT').count()
        locked_marks = submissions.filter(is_locked=True).count()

        grades_uploaded = FinalGrade.objects.filter(course_offering=offering).count()

        frozen_results = FinalGrade.objects.filter(
            course_offering=offering,
            is_frozen=True
        ).count()

        dashboard_data.append({
            "offering": offering,
            "total_students": total_students,
            "locked_marks": locked_marks,
            "grades_uploaded": grades_uploaded,
            "frozen_results": frozen_results,
        })

    return render(request, "exam/dashboard.html", {
        "dashboard_data": dashboard_data,
        "total_courses": offerings.count(),
        "locked_courses": sum(
            1 for d in dashboard_data
            if d["locked_marks"] == d["total_students"] and d["total_students"] > 0
        ),
        "uploaded_grades": sum(1 for d in dashboard_data if d["grades_uploaded"] > 0),
        "frozen_results": sum(1 for d in dashboard_data if d["frozen_results"] > 0),
    })


# =========================================================
# 📥 DOWNLOAD TEMPLATE
# =========================================================
@login_required
def download_grades_template(request, offering_id):

    if not (is_exam_section(request.user) or is_faculty(request.user)):
        return HttpResponseForbidden()

    offering = get_object_or_404(CourseOffering, id=offering_id)

    students = User.objects.filter(
        enrollment__offering=offering,
        enrollment__is_active=True
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Roll No", "Minor1", "Minor2", "Mid", "End"])

    for student in students:
        ws.append([student.username, "", "", "", ""])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="marks_template.xlsx"'

    wb.save(response)
    return response

# =========================================================
# 📊 SEMESTER RESULT UPLOAD (SGPA / CGPA)
# =========================================================
@login_required
def exam_upload_semester_results(request, semester_id):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    semester_results = SemesterResult.objects.filter(semester_id=semester_id)

    if semester_results.filter(is_finalized=True).exists():
        return HttpResponseForbidden("Results already finalized")

    if request.method == "POST":

        file = request.FILES.get("file")

        if not file:
            return render(request, "exam/upload_semester_results.html", {
                "error": "File required",
                "semester_id": semester_id
            })

        try:
            wb = openpyxl.load_workbook(file)
        except:
            return render(request, "exam/upload_semester_results.html", {
                "error": "Invalid Excel file",
                "semester_id": semester_id
            })

        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        expected = ["Roll No", "SGPA", "CGPA"]

        if headers != expected:
            return render(request, "exam/upload_semester_results.html", {
                "error": f"Expected {expected}",
                "semester_id": semester_id
            })

        errors = []
        processed = set()

        with transaction.atomic():
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

                roll_no, sgpa, cgpa = row

                roll_no = str(roll_no).strip()

                try:
                    student = User.objects.filter(username__iexact=roll_no).first()

                    if not student:
                        errors.append(f"Row {row_index}: Student not found ({roll_no})")
                        continue
                except User.DoesNotExist:
                    errors.append(f"Row {row_index}: Student not found")
                    continue

                try:
                    sgpa = float(sgpa)
                    cgpa = float(cgpa)
                except:
                    errors.append(f"Row {row_index}: Invalid SGPA/CGPA")
                    continue

                processed.add(student.id)

                SemesterResult.objects.update_or_create(
                    student=student,
                    semester_id=semester_id,
                    defaults={
                        "sgpa": sgpa,
                        "cgpa": cgpa,
                        "published_by": request.user
                    }
                )

        if errors:
            return render(request, "exam/upload_semester_results.html", {
                "error": errors,
                "semester_id": semester_id
            })

        return redirect("dashboard")

    return render(request, "exam/upload_semester_results.html", {
        "semester_id": semester_id
    })

# =========================================================
# 🎓 STUDENT VIEW GRADES
# =========================================================
@login_required
def student_view_grades(request):

    if not is_student(request.user):
        return HttpResponseForbidden()

    student = request.user

    grades = FinalGrade.objects.filter(
        student=student,
        is_published=True
    ).select_related("course_offering", "course_offering__course")

    semesters = Semester.objects.all()

    semester_results = []

    for sem in semesters:
        sgpa = calculate_sgpa(student, sem)

        if sgpa > 0:
            semester_results.append({
                "semester": sem,
                "sgpa": sgpa
            })

    cgpa = calculate_cgpa(student)

    return render(request, "student/view_grades.html", {
        "grades": grades,
        "semester_results": semester_results,
        "cgpa": cgpa
    })
# =========================================================
# 📥 DOWNLOAD FINAL GRADES TEMPLATE (EXAM)
# =========================================================
@login_required
def download_final_grades_template(request, offering_id):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    offering = get_object_or_404(CourseOffering, id=offering_id)

    students = User.objects.filter(
        enrollment__offering=offering,
        enrollment__is_active=True
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    # ✅ Correct headers for FINAL GRADES
    ws.append(["Roll No", "Name", "Course Code", "Course Name", "Grade"])

    for student in students:
        ws.append([
            student.username,
            student.get_full_name(),
            offering.course.course_code,
            offering.course.course_title,
            ""  # grade to fill
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="final_grades_template.xlsx"'

    wb.save(response)
    return response

# =========================================================
# 📢 RELEASE RESULTS
# =========================================================
@login_required
def exam_release_results(request, offering_id):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    offering = get_object_or_404(CourseOffering, id=offering_id)

    grades = FinalGrade.objects.filter(course_offering=offering)

    if not grades.exists():
        return HttpResponseForbidden("No grades uploaded")

    if request.method == "POST":
        grades.update(is_published=True)
        return redirect("dashboard")

    return render(request, "exam/confirm_release.html", {
        "offering": offering
    })

@login_required
def exam_release_semester_results(request, semester_id):
    print("User:", request.user)
    print("Role:", getattr(request.user, "role", None))
    offerings = CourseOffering.objects.filter(semester_id=semester_id)

    for o in offerings:

        total = MarksSubmission.objects.filter(course_offering=o).count()
        graded = FinalGrade.objects.filter(course_offering=o).count()
        frozen = FinalGrade.objects.filter(course_offering=o, is_frozen=True).exists()

        print("------")
        print("Course:", o.course.course_title)
        print("Total students:", total)
        print("Graded:", graded)
        print("Frozen:", frozen)

        if total != graded:
            return HttpResponseForbidden(f"{o.course.course_title} → not fully graded")

        if not frozen:
            return HttpResponseForbidden(f"{o.course.course_title} → not frozen")
    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    offerings = CourseOffering.objects.filter(semester_id=semester_id)

    submissions = MarksSubmission.objects.filter(course_offering__in=offerings)
    grades = FinalGrade.objects.filter(course_offering__in=offerings)

    total_students = MarksSubmission.objects.filter(
        course_offering__in=offerings
    ).count()
    total_grades = FinalGrade.objects.filter(
    course_offering__in=offerings
).count()

    # ❌ Block if incomplete
    if total_students == 0 or total_grades != total_students:
        return HttpResponseForbidden(
            f"Incomplete: {total_grades}/{total_students} grades uploaded"
        )

    # ❌ Block if not frozen fully
    if grades.filter(is_frozen=False).exists():
        return HttpResponseForbidden("Freeze all courses before release")

    if request.method == "POST":
        grades.update(is_published=True)
        return redirect("exam_dashboard")

    return render(request, "exam/confirm_release_semester.html", {
        "semester_id": semester_id
    })

@login_required
def year_semester_view(request, year):

    if not is_exam_section(request.user):
        return HttpResponseForbidden()

    semesters_map = {
        1: [1,2],
        2: [3,4],
        3: [5,6],
        4: [7,8]
    }

    offerings = CourseOffering.objects.filter(
        semester__year=year
    ).select_related("course", "semester")

    data = []

    for offering in offerings:

        submissions = MarksSubmission.objects.filter(course_offering=offering)
        grades = FinalGrade.objects.filter(course_offering=offering)

        total_students = submissions.count()
        locked = submissions.filter(is_locked=True).count()

        grades_uploaded = grades.count()
        frozen = grades.filter(is_frozen=True).count()

        data.append({
            "offering": offering,
            "total_students": total_students,
            "locked": locked,
            "grades_uploaded": grades_uploaded,
            "frozen": frozen
        })

    return render(request, "exam/year_semester.html", {
        "year": year,
        "data": data
    })

def calculate_sgpa(student, semester):

    grades = FinalGrade.objects.filter(
        student=student,
        course_offering__semester=semester,
        is_published=True
    )

    total_points = 0
    total_credits = 0

    for g in grades:
        credits = g.course.credits
        gp = GRADE_POINTS.get(g.grade_letter, 0)

        total_points += gp * credits
        total_credits += credits

    if total_credits == 0:
        return 0

    return round(total_points / total_credits, 2)

def calculate_cgpa(student):

    grades = FinalGrade.objects.filter(
        student=student,
        is_published=True
    )

    best_courses = {}

    # 🔥 take BEST attempt per course
    for g in grades:
        key = g.course.id

        if key not in best_courses:
            best_courses[key] = g
        else:
            if GRADE_POINTS[g.grade_letter] > GRADE_POINTS[best_courses[key].grade]:
                best_courses[key] = g

    total_points = 0
    total_credits = 0

    for g in best_courses.values():
        credits = g.course.credits
        gp = GRADE_POINTS[g.grade_letter]

        total_points += gp * credits
        total_credits += credits

    if total_credits == 0:
        return 0

    return round(total_points / total_credits, 2)

def get_backlogs(student):

    grades = FinalGrade.objects.filter(student=student)

    backlog_courses = {}

    for g in grades:
        key = g.course.id

        if key not in backlog_courses:
            backlog_courses[key] = g
        else:
            if GRADE_POINTS[g.grade_letter] > GRADE_POINTS[backlog_courses[key].grade]:
                backlog_courses[key] = g

    # only failed ones
    return [
        g for g in backlog_courses.values()
        if g.grade_letter == "F"
    ]

@login_required
def student_result_dashboard(request):

    student = request.user

    semesters = Semester.objects.all()

    semester_results = []

    for sem in semesters:

        sgpa = calculate_sgpa(student, sem)

        semester_results.append({
            "semester": sem,
            "sgpa": sgpa
        })

    cgpa = calculate_cgpa(student)
    backlogs = get_backlogs(student)

    return render(request, "student/results.html", {
        "semester_results": semester_results,
        "cgpa": cgpa,
        "backlogs": backlogs
    })